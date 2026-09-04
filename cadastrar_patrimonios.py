import csv
import getpass
import math
import os
import sys
import time
import unicodedata
from datetime import date, datetime

import openpyxl
import requests


# ============================================================
# CONFIGURAÇÕES
# ============================================================

ARQUIVO_PLANILHA = "Patrimonios_Organizado.xlsx"

BASE_URL = "http://127.0.0.1:5000"

URL_LOGIN = f"{BASE_URL}/login"
URL_CADASTRO = f"{BASE_URL}/api/equipamentos"
URL_BOOTSTRAP = f"{BASE_URL}/api/bootstrap"
URL_LIXEIRA = f"{BASE_URL}/api/lixeira"

CAMPO_USUARIO = "email"
CAMPO_SENHA = "senha"

# Se True, cadastra somente o primeiro registro para teste.
# Depois de confirmar que funcionou, altere para False.
TESTAR_PRIMEIRO = False

PAUSA = 0.3

# Ignora registros cujo patrimônio já esteja cadastrado.
IGNORAR_DUPLICADOS = True

# Categoria padrão quando não for possível identificar automaticamente.
CATEGORIA_PADRAO = "Informática"

# Motivo utilizado quando a planilha possui data de baixa.
MOTIVO_BAIXA_PADRAO = (
    "Equipamento baixado conforme planilha de patrimônio"
)

# Nome das colunas esperadas/principais.
COLUNAS = {
    "patrimonio": "Patrimônio",
    "equipamento": "Equipamento",
    "marca": "Marca",
    "modelo": "Modelo",
    "local_setor": "Local/Setor",
    "responsavel": "Responsável",
    "data_aquisicao": "Data de Aquisição",
    "data_baixa": "Data da Baixa",
    "observacoes": "Observações",
}


# ============================================================
# ALIASES DE CABEÇALHOS
# ============================================================

ALIASES = {
    "patrimonio": [
        "patrimonio",
        "patrimonioid",
        "codigo",
        "codigopatrimonio",
        "codpatrimonio",
        "numeropatrimonio",
        "numero patrimonio",
        "n patrimonio",
        "tombamento",
        "numero tombamento",
    ],

    "equipamento": [
        "equipamento",
        "nome",
        "nome equipamento",
        "descricao",
        "descricao equipamento",
        "bem",
        "item",
        "produto",
    ],

    "marca": [
        "marca",
        "fabricante",
        "marca fabricante",
        "fabricante marca",
    ],

    "modelo": [
        "modelo",
        "modelo equipamento",
    ],

    "local_setor": [
        "local setor",
        "local/setor",
        "local",
        "setor",
        "setor local",
        "unidade",
        "departamento",
        "lotacao",
        "locacao",
        "localizacao",
        "localizacao equipamento",
    ],

    "responsavel": [
        "responsavel",
        "responsavel equipamento",
        "responsavel pelo equipamento",
        "responsavel usuario",
        "usuario",
        "usuario responsavel",
        "servidor responsavel",
    ],

    "data_aquisicao": [
        "data aquisicao",
        "aquisicao",
        "data de aquisicao",
        "data entrada",
        "data de entrada",
        "data compra",
        "data de compra",
    ],

    "data_baixa": [
        "data baixa",
        "data de baixa",
        "baixa",
        "data descarte",
        "data de descarte",
        "data baixa patrimonio",
    ],

    "observacoes": [
        "observacoes",
        "observacao",
        "obs",
        "observacao geral",
        "observacoes gerais",
        "comentarios",
        "comentario",
        "informacoes",
        "informacao",
    ],
}


# ============================================================
# NORMALIZAÇÃO
# ============================================================

def remover_acentos(texto):
    """Remove acentos de uma string."""
    texto = str(texto)

    normalizado = unicodedata.normalize(
        "NFD",
        texto
    )

    return "".join(
        caractere
        for caractere in normalizado
        if unicodedata.category(caractere) != "Mn"
    )


def normalizar_cabecalho(texto):
    """
    Normaliza cabeçalhos para permitir diferenças como:

    Responsável
    RESPONSÁVEL
    responsavel
    Responsável pelo equipamento
    """
    if texto is None:
        return ""

    texto = str(texto)

    texto = texto.replace("\xa0", " ")
    texto = remover_acentos(texto)

    texto = texto.lower().strip()

    # Padroniza alguns caracteres.
    texto = texto.replace("º", "")
    texto = texto.replace("°", "")
    texto = texto.replace("&", "e")

    # Mantém somente letras e números.
    texto = "".join(
        caractere if caractere.isalnum() else " "
        for caractere in texto
    )

    # Remove espaços duplicados.
    texto = " ".join(texto.split())

    return texto


def chave_cabecalho(texto):
    """
    Remove espaços para facilitar a comparação.
    """
    return normalizar_cabecalho(texto).replace(" ", "")


def valor_vazio(valor):
    """
    Verifica se o valor é realmente vazio.

    IMPORTANTE:
    'não informado' NÃO é considerado vazio.
    """
    if valor is None:
        return True

    if isinstance(valor, str):
        texto = valor.replace("\xa0", " ").strip()

        if texto == "":
            return True

    return False


# ============================================================
# LIMPEZA DE VALORES
# ============================================================

def limpar(valor, manter_nao_informado=True):
    """
    Converte valores da planilha para texto.

    Diferentemente do importador anterior,
    'não informado' é preservado.
    """

    if valor is None:
        return ""

    # Datas do Excel.
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")

    # Valores numéricos.
    if isinstance(valor, float):
        if math.isnan(valor):
            return ""

        if valor.is_integer():
            return str(int(valor))

        return str(valor).strip()

    if isinstance(valor, int):
        return str(valor)

    texto = str(valor)

    texto = texto.replace("\xa0", " ")
    texto = texto.strip()

    if texto == "":
        return ""

    # Alguns valores que realmente representam vazio.
    vazios = {
        "",
        "nan",
        "NaN",
        "NONE",
        "None",
        "NULL",
        "Null",
        "N/A",
        "n/a",
    }

    if texto in vazios:
        return ""

    # NÃO transformar "não informado" em vazio.
    if manter_nao_informado:
        return texto

    return texto


# ============================================================
# DATAS
# ============================================================

def converter_data(valor):
    """
    Converte datas do Excel para YYYY-MM-DD.
    """

    if valor is None:
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, (int, float)):
        try:
            if isinstance(valor, float) and math.isnan(valor):
                return ""

            # Datas do Excel geralmente são números.
            # O openpyxl normalmente já converte células
            # formatadas como data, mas deixamos este suporte.
            if 1 <= float(valor) <= 100000:
                from openpyxl.utils.datetime import from_excel

                data = from_excel(valor)

                if isinstance(data, datetime):
                    return data.strftime("%Y-%m-%d")

                if isinstance(data, date):
                    return data.strftime("%Y-%m-%d")
        except Exception:
            pass

    texto = limpar(valor)

    if not texto:
        return ""

    formatos = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%Y/%m/%d",
        "%d/%m/%y",
        "%d-%m-%y",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(
                texto,
                formato
            ).strftime("%Y-%m-%d")
        except ValueError:
            pass

    return texto


# ============================================================
# CATEGORIA
# ============================================================

def descobrir_categoria(nome, marca="", modelo=""):
    """
    Tenta identificar automaticamente a categoria.
    """

    texto = f"{nome} {marca} {modelo}".lower()

    categorias = [
        (
            [
                "notebook",
                "laptop",
                "computador",
                "desktop",
                "pc",
                "monitor",
                "impressora",
                "scanner",
                "projetor",
                "datashow",
                "estabilizador",
                "nobreak",
                "switch",
                "roteador",
                "access point",
                "tablet",
                "celular",
                "telefone",
                "telefone ip",
                "servidor",
                "teclado",
                "mouse",
                "webcam",
                "headset",
                "caixa de som",
                "som",
                "microfone",
                "hd",
                "ssd",
                "pendrive",
                "computador",
            ],
            "Informática",
        ),

        (
            [
                "mesa",
                "cadeira",
                "armario",
                "armário",
                "arquivo",
                "estante",
                "balcao",
                "balcão",
                "gaveteiro",
                "mobiliario",
                "mobiliário",
            ],
            "Mobiliário",
        ),

        (
            [
                "ar condicionado",
                "condicionador de ar",
                "ventilador",
                "geladeira",
                "refrigerador",
                "microondas",
                "micro-ondas",
                "fogao",
                "fogão",
                "forno",
                "bebedouro",
                "purificador",
            ],
            "Eletrodomésticos",
        ),

        (
            [
                "projetor",
                "data show",
                "datashow",
            ],
            "Informática",
        ),
    ]

    for palavras, categoria in categorias:
        for palavra in palavras:
            if palavra in texto:
                return categoria

    return CATEGORIA_PADRAO


# ============================================================
# CÉLULAS MESCLADAS
# ============================================================

def construir_mapa_mescladas(ws, colunas_interesse):
    """
    Cria um mapa para células mescladas.

    Exemplo:

    Marca:
    Dell
    [célula mesclada]
    [célula mesclada]

    Visualmente o Excel mostra Dell nas três linhas,
    mas o openpyxl normalmente só encontra o valor na
    primeira célula.

    Esta função corrige isso.
    """

    mapa = {}

    if not ws.merged_cells.ranges:
        return mapa

    colunas_interesse = set(colunas_interesse)

    for intervalo in ws.merged_cells.ranges:

        min_col = intervalo.min_col
        max_col = intervalo.max_col
        min_row = intervalo.min_row
        max_row = intervalo.max_row

        # Ignora mesclagens que não atingem as colunas usadas.
        if not any(
            min_col <= coluna <= max_col
            for coluna in colunas_interesse
        ):
            continue

        # Valor verdadeiro fica no canto superior esquerdo.
        valor = ws.cell(
            row=min_row,
            column=min_col
        ).value

        for linha in range(min_row, max_row + 1):
            for coluna in range(min_col, max_col + 1):

                if coluna in colunas_interesse:
                    mapa[(linha, coluna)] = valor

    return mapa


# ============================================================
# DESCOBERTA DO CABEÇALHO
# ============================================================

def descobrir_linha_cabecalho(ws):
    """
    Procura automaticamente a linha dos cabeçalhos.

    Isso evita problemas caso a planilha tenha título,
    logotipo ou outras linhas antes da tabela.
    """

    melhores = None

    limite = min(ws.max_row, 30)

    nomes_validos = set()

    for aliases in ALIASES.values():
        for alias in aliases:
            nomes_validos.add(
                chave_cabecalho(alias)
            )

    for numero_linha in range(1, limite + 1):

        encontrados = 0
        campos = set()

        for coluna in range(1, ws.max_column + 1):

            valor = ws.cell(
                row=numero_linha,
                column=coluna
            ).value

            chave = chave_cabecalho(valor)

            if not chave:
                continue

            for campo, aliases in ALIASES.items():

                aliases_normalizados = {
                    chave_cabecalho(alias)
                    for alias in aliases
                }

                if chave in aliases_normalizados:
                    encontrados += 1
                    campos.add(campo)
                    break

        if melhores is None or len(campos) > len(melhores[1]):
            melhores = (
                numero_linha,
                campos,
                encontrados
            )

    if melhores is None:
        raise RuntimeError(
            "Não foi possível localizar a linha dos cabeçalhos."
        )

    linha, campos, encontrados = melhores

    if len(campos) < 3:
        raise RuntimeError(
            "Não foi possível identificar os cabeçalhos da planilha."
        )

    return linha


# ============================================================
# MAPEAMENTO DAS COLUNAS
# ============================================================

def mapear_colunas(ws, linha_cabecalho):
    """
    Relaciona os nomes das colunas do Excel aos campos
    utilizados pelo sistema.
    """

    mapa = {}

    for coluna in range(1, ws.max_column + 1):

        valor = ws.cell(
            row=linha_cabecalho,
            column=coluna
        ).value

        chave = chave_cabecalho(valor)

        if not chave:
            continue

        for campo, aliases in ALIASES.items():

            aliases_normalizados = {
                chave_cabecalho(alias)
                for alias in aliases
            }

            if chave in aliases_normalizados:

                # Não sobrescrever se já encontrou.
                if campo not in mapa:
                    mapa[campo] = coluna

                break

    # Campos realmente obrigatórios na origem.
    obrigatorios = [
        "patrimonio",
        "equipamento",
        "marca",
        "local_setor",
        "responsavel",
    ]

    faltando = [
        campo
        for campo in obrigatorios
        if campo not in mapa
    ]

    if faltando:
        nomes = [
            COLUNAS[campo]
            for campo in faltando
        ]

        raise RuntimeError(
            "Colunas obrigatórias não encontradas: "
            + ", ".join(nomes)
        )

    return mapa


# ============================================================
# LEITURA DE CÉLULA
# ============================================================

def obter_valor(ws, linha, coluna, mapa_mescladas):
    """
    Lê uma célula normalmente.

    Se ela fizer parte de uma célula mesclada,
    recupera o valor da célula superior esquerda.
    """

    chave = (linha, coluna)

    if chave in mapa_mescladas:
        valor = mapa_mescladas[chave]
    else:
        valor = ws.cell(
            row=linha,
            column=coluna
        ).value

    return limpar(valor)


# ============================================================
# LEITURA DA PLANILHA
# ============================================================

def ler_planilha(caminho):
    """
    Lê a planilha e retorna uma lista de equipamentos.
    """

    if not os.path.exists(caminho):
        raise FileNotFoundError(
            f"Planilha não encontrada: {caminho}"
        )

    print()
    print("=" * 70)
    print("LENDO PLANILHA")
    print("=" * 70)
    print(f"Arquivo: {os.path.abspath(caminho)}")

    # Carregamos uma vez para valores.
    wb = openpyxl.load_workbook(
        caminho,
        data_only=True
    )

    ws = wb.active

    print(f"Aba utilizada: {ws.title}")
    print(f"Linhas encontradas: {ws.max_row}")
    print(f"Colunas encontradas: {ws.max_column}")

    linha_cabecalho = descobrir_linha_cabecalho(ws)

    print(
        f"Linha do cabeçalho identificada: "
        f"{linha_cabecalho}"
    )

    mapa_colunas = mapear_colunas(
        ws,
        linha_cabecalho
    )

    print()
    print("COLUNAS IDENTIFICADAS:")

    for campo, coluna in mapa_colunas.items():

        nome = ws.cell(
            row=linha_cabecalho,
            column=coluna
        ).value

        print(
            f"  {campo:20} -> coluna {coluna}: {nome}"
        )

    # Colunas usadas na leitura.
    colunas_interesse = list(
        mapa_colunas.values()
    )

    mapa_mescladas = construir_mapa_mescladas(
        ws,
        colunas_interesse
    )

    if mapa_mescladas:
        print(
            f"\nCélulas mescladas tratadas: "
            f"{len(mapa_mescladas)}"
        )

    registros = []

    estatisticas = {
        "patrimonio": 0,
        "equipamento": 0,
        "marca": 0,
        "local_setor": 0,
        "responsavel": 0,
    }

    for linha in range(
        linha_cabecalho + 1,
        ws.max_row + 1
    ):

        def campo(nome):
            coluna = mapa_colunas.get(nome)

            if coluna is None:
                return ""

            return obter_valor(
                ws,
                linha,
                coluna,
                mapa_mescladas
            )

        patrimonio = campo("patrimonio")
        equipamento = campo("equipamento")
        marca = campo("marca")
        modelo = campo("modelo")
        local = campo("local_setor")
        responsavel = campo("responsavel")
        data_aquisicao = converter_data(
            campo("data_aquisicao")
        )
        data_baixa = converter_data(
            campo("data_baixa")
        )
        observacoes = campo("observacoes")

        # Ignora linhas totalmente vazias.
        valores_principais = [
            patrimonio,
            equipamento,
            marca,
            modelo,
            local,
            responsavel,
            data_aquisicao,
            data_baixa,
            observacoes,
        ]

        if all(
            valor_vazio(valor)
            for valor in valores_principais
        ):
            continue

        # Estatísticas.
        for campo_estatistica in estatisticas:
            valor = locals().get(
                {
                    "patrimonio": "patrimonio",
                    "equipamento": "equipamento",
                    "marca": "marca",
                    "local_setor": "local",
                    "responsavel": "responsavel",
                }[campo_estatistica]
            )

            if valor_vazio(valor):
                estatisticas[campo_estatistica] += 1

        categoria = descobrir_categoria(
            equipamento,
            marca,
            modelo
        )

        item = {
            "barcode": patrimonio,
            "nome": equipamento,
            "marca": marca,
            "modelo": modelo,
            "serie": "",
            "categoria": categoria,
            "local": local,
            "responsavel": responsavel,
            "dataAquisicao": data_aquisicao,
            "valor": "",
            "observacoes": observacoes,
            "_linha": linha,
            "_data_baixa": data_baixa,
        }

        registros.append(item)

    wb.close()

    print()
    print("=" * 70)
    print("RESUMO DA LEITURA")
    print("=" * 70)

    print(
        f"Registros encontrados: {len(registros)}"
    )

    for campo, quantidade in estatisticas.items():

        if quantidade:
            print(
                f"Linhas com {campo} vazio: "
                f"{quantidade}"
            )

    print()
    print(
        "IMPORTANTE: 'não informado' foi preservado "
        "como texto e NÃO será tratado como campo vazio."
    )

    return registros


# ============================================================
# VALIDAÇÃO
# ============================================================

def validar_item(item):
    """
    Valida os campos que são obrigatórios no app.py.
    """

    obrigatorios = {
        "barcode": "Patrimônio",
        "nome": "Equipamento",
        "marca": "Marca",
        "categoria": "Categoria",
        "local": "Local/Setor",
        "responsavel": "Responsável",
    }

    vazios = []

    for campo, nome in obrigatorios.items():

        valor = item.get(campo, "")

        if valor_vazio(valor):
            vazios.append(nome)

    if vazios:
        return False, (
            "Campos obrigatórios vazios: "
            + ", ".join(vazios)
        )

    return True, ""


# ============================================================
# LOGIN
# ============================================================

def fazer_login(sessao):
    """
    Faz login no Flask.
    """

    print()
    print("=" * 70)
    print("LOGIN")
    print("=" * 70)

    email = input("E-mail: ").strip()

    senha = getpass.getpass(
        "Senha: "
    )

    try:

        resposta = sessao.post(
            URL_LOGIN,
            data={
                CAMPO_USUARIO: email,
                CAMPO_SENHA: senha,
            },
            timeout=30,
            allow_redirects=True,
        )

    except requests.RequestException as erro:

        raise RuntimeError(
            "Não foi possível conectar ao sistema.\n"
            f"Verifique se o Flask está executando em:\n"
            f"{BASE_URL}\n\n"
            f"Erro: {erro}"
        )

    if resposta.status_code != 200:

        raise RuntimeError(
            f"Falha no login. HTTP {resposta.status_code}"
        )

    # Verifica se a sessão realmente está autenticada.
    try:

        teste = sessao.get(
            URL_BOOTSTRAP,
            timeout=30,
        )

        if teste.status_code != 200:

            raise RuntimeError(
                "Login não foi validado pelo sistema."
            )

    except requests.RequestException as erro:

        raise RuntimeError(
            f"Erro ao verificar sessão: {erro}"
        )

    print()
    print("LOGIN REALIZADO COM SUCESSO.")

    return True


# ============================================================
# ERRO DA API
# ============================================================

def mensagem_erro(resposta):
    """
    Tenta extrair uma mensagem amigável da API.
    """

    try:

        dados = resposta.json()

        if isinstance(dados, dict):

            erro = dados.get("error")

            if erro:
                return str(erro)

            mensagem = dados.get("message")

            if mensagem:
                return str(mensagem)

    except Exception:
        pass

    texto = resposta.text.strip()

    if texto:
        # Evita mostrar uma página HTML enorme.
        if len(texto) > 500:
            texto = texto[:500] + "..."

        return texto

    return (
        f"HTTP {resposta.status_code}"
    )


# ============================================================
# CARREGAR EQUIPAMENTOS EXISTENTES
# ============================================================

def carregar_existentes(sessao):
    """
    Carrega equipamentos ativos já existentes.
    """

    print()
    print("Consultando equipamentos já cadastrados...")

    resposta = sessao.get(
        URL_BOOTSTRAP,
        timeout=60,
    )

    if resposta.status_code != 200:

        raise RuntimeError(
            "Não foi possível consultar os equipamentos "
            "existentes.\n"
            + mensagem_erro(resposta)
        )

    dados = resposta.json()

    equipamentos = dados.get(
        "equipamentos",
        []
    )

    existentes = {}

    for equipamento in equipamentos:

        barcode = limpar(
            equipamento.get("barcode")
        )

        if barcode:
            existentes[barcode] = equipamento

    print(
        f"Equipamentos ativos encontrados: "
        f"{len(existentes)}"
    )

    return existentes


# ============================================================
# CARREGAR LIXEIRA
# ============================================================

def carregar_lixeira(sessao):
    """
    Carrega patrimônios que estão na lixeira.

    Isso é importante porque o banco possui UNIQUE no barcode.
    Mesmo que um equipamento esteja excluído logicamente,
    o SQLite continua impedindo outro registro com o mesmo
    patrimônio.
    """

    print(
        "Consultando equipamentos na lixeira..."
    )

    resposta = sessao.get(
        URL_LIXEIRA,
        timeout=60,
    )

    if resposta.status_code != 200:

        print(
            "Aviso: não foi possível consultar a lixeira."
        )

        return {}

    try:

        dados = resposta.json()

    except Exception:

        return {}

    equipamentos = dados.get(
        "equipamentos",
        []
    )

    lixeira = {}

    for equipamento in equipamentos:

        barcode = limpar(
            equipamento.get("barcode")
        )

        if barcode:
            lixeira[barcode] = equipamento

    print(
        f"Equipamentos na lixeira: "
        f"{len(lixeira)}"
    )

    return lixeira


# ============================================================
# CADASTRAR EQUIPAMENTO
# ============================================================

def cadastrar_equipamento(sessao, item):
    """
    Envia o equipamento para:

    POST /api/equipamentos
    """

    payload = {
        "barcode": item["barcode"],
        "nome": item["nome"],
        "marca": item["marca"],
        "modelo": item.get("modelo", ""),
        "serie": item.get("serie", ""),
        "categoria": item["categoria"],
        "local": item["local"],
        "responsavel": item["responsavel"],
        "dataAquisicao": item.get(
            "dataAquisicao",
            ""
        ),
        "valor": item.get(
            "valor",
            ""
        ),
        "observacoes": item.get(
            "observacoes",
            ""
        ),
    }

    try:

        resposta = sessao.post(
            URL_CADASTRO,
            json=payload,
            timeout=60,
        )

    except requests.RequestException as erro:

        return False, (
            "Erro de conexão: "
            f"{erro}"
        ), None

    if resposta.status_code in (200, 201):

        try:
            equipamento = resposta.json()
        except Exception:
            equipamento = None

        return True, "", equipamento

    return False, mensagem_erro(resposta), None


# ============================================================
# DAR BAIXA
# ============================================================

def dar_baixa(sessao, item, equipamento):
    """
    Dá baixa em um equipamento quando a planilha possui
    uma data de baixa.
    """

    data_baixa = item.get(
        "_data_baixa",
        ""
    )

    if not data_baixa:
        return True, ""

    equipamento_id = equipamento.get("id")

    if not equipamento_id:
        return False, (
            "Equipamento cadastrado, mas a API não "
            "retornou o ID para dar baixa."
        )

    payload = {
        "motivo": MOTIVO_BAIXA_PADRAO,
        "data": data_baixa,
        "responsavel": item["responsavel"],
        "observacoes": item.get(
            "observacoes",
            ""
        ),
    }

    try:

        resposta = sessao.post(
            f"{BASE_URL}/api/equipamentos/"
            f"{equipamento_id}/baixa",
            json=payload,
            timeout=60,
        )

    except requests.RequestException as erro:

        return False, (
            "Erro de conexão ao dar baixa: "
            f"{erro}"
        )

    if resposta.status_code in (200, 201):

        return True, ""

    return False, mensagem_erro(resposta)


# ============================================================
# CSV DE FALHAS
# ============================================================

def salvar_falhas(falhas, arquivo="falhas_cadastro.csv"):
    """
    Salva somente os registros que falharam.
    """

    if not falhas:
        return

    campos = [
        "linha",
        "patrimonio",
        "equipamento",
        "marca",
        "modelo",
        "local",
        "responsavel",
        "categoria",
        "data_aquisicao",
        "data_baixa",
        "motivo",
    ]

    try:

        with open(
            arquivo,
            "w",
            newline="",
            encoding="utf-8-sig",
        ) as arquivo_csv:

            escritor = csv.DictWriter(
                arquivo_csv,
                fieldnames=campos,
            )

            escritor.writeheader()

            for falha in falhas:

                escritor.writerow({
                    campo: falha.get(
                        campo,
                        ""
                    )
                    for campo in campos
                })

        print()
        print(
            f"Arquivo de falhas salvo em:\n"
            f"{os.path.abspath(arquivo)}"
        )

    except Exception as erro:

        print(
            f"Não foi possível salvar o CSV de falhas: "
            f"{erro}"
        )


# ============================================================
# EXIBIR ITEM
# ============================================================

def mostrar_item(item):
    """
    Mostra os dados que serão enviados.
    """

    print()
    print("-" * 70)
    print(
        f"Linha Excel: {item.get('_linha')}"
    )
    print(
        f"Patrimônio : [{item.get('barcode', '')}]"
    )
    print(
        f"Equipamento: [{item.get('nome', '')}]"
    )
    print(
        f"Marca      : [{item.get('marca', '')}]"
    )
    print(
        f"Modelo     : [{item.get('modelo', '')}]"
    )
    print(
        f"Categoria  : [{item.get('categoria', '')}]"
    )
    print(
        f"Local      : [{item.get('local', '')}]"
    )
    print(
        f"Responsável: [{item.get('responsavel', '')}]"
    )
    print(
        f"Data aq.   : [{item.get('dataAquisicao', '')}]"
    )
    print(
        f"Data baixa : [{item.get('_data_baixa', '')}]"
    )
    print("-" * 70)


# ============================================================
# MAIN
# ============================================================

def main():

    print()
    print("=" * 70)
    print("IMPORTADOR DE PATRIMÔNIOS")
    print("=" * 70)
    print()
    print(
        f"Servidor: {BASE_URL}"
    )
    print(
        f"Planilha : {os.path.abspath(ARQUIVO_PLANILHA)}"
    )

    # --------------------------------------------------------
    # 1. Ler planilha
    # --------------------------------------------------------

    try:

        registros = ler_planilha(
            ARQUIVO_PLANILHA
        )

    except Exception as erro:

        print()
        print("=" * 70)
        print("ERRO AO LER A PLANILHA")
        print("=" * 70)
        print(str(erro))
        print()

        input(
            "Pressione ENTER para sair..."
        )

        return

    if not registros:

        print()
        print(
            "Nenhum registro encontrado."
        )

        input(
            "Pressione ENTER para sair..."
        )

        return

    # --------------------------------------------------------
    # 2. Login
    # --------------------------------------------------------

    sessao = requests.Session()

    try:

        fazer_login(sessao)

    except Exception as erro:

        print()
        print("=" * 70)
        print("ERRO NO LOGIN")
        print("=" * 70)
        print(str(erro))
        print()

        input(
            "Pressione ENTER para sair..."
        )

        return

    # --------------------------------------------------------
    # 3. Consultar banco
    # --------------------------------------------------------

    try:

        existentes = carregar_existentes(
            sessao
        )

        lixeira = carregar_lixeira(
            sessao
        )

    except Exception as erro:

        print()
        print("=" * 70)
        print("ERRO AO CONSULTAR O SISTEMA")
        print("=" * 70)
        print(str(erro))
        print()

        input(
            "Pressione ENTER para sair..."
        )

        return

    # --------------------------------------------------------
    # 4. Teste
    # --------------------------------------------------------

    registros_processar = registros

    if TESTAR_PRIMEIRO:

        registros_processar = registros[:1]

        print()
        print(
            "MODO TESTE ATIVADO."
        )
        print(
            "Somente o primeiro registro será processado."
        )

    # --------------------------------------------------------
    # 5. Processamento
    # --------------------------------------------------------

    sucessos = 0
    duplicados = 0
    na_lixeira = 0
    falhas = 0
    baixas = 0

    lista_falhas = []

    total = len(registros_processar)

    print()
    print("=" * 70)
    print(
        f"INICIANDO IMPORTAÇÃO DE {total} REGISTROS"
    )
    print("=" * 70)

    for indice, item in enumerate(
        registros_processar,
        start=1
    ):

        linha = item.get(
            "_linha",
            "?"
        )

        patrimonio = limpar(
            item.get("barcode")
        )

        print()
        print(
            f"[{indice}/{total}] "
            f"Linha Excel {linha} "
            f"| Patrimônio: {patrimonio}"
        )

        # ----------------------------------------------------
        # Validação local
        # ----------------------------------------------------

        valido, motivo = validar_item(
            item
        )

        if not valido:

            falhas += 1

            print(
                f"  FALHOU - {motivo}"
            )

            mostrar_item(item)

            lista_falhas.append({
                "linha": linha,
                "patrimonio": item.get(
                    "barcode",
                    ""
                ),
                "equipamento": item.get(
                    "nome",
                    ""
                ),
                "marca": item.get(
                    "marca",
                    ""
                ),
                "modelo": item.get(
                    "modelo",
                    ""
                ),
                "local": item.get(
                    "local",
                    ""
                ),
                "responsavel": item.get(
                    "responsavel",
                    ""
                ),
                "categoria": item.get(
                    "categoria",
                    ""
                ),
                "data_aquisicao": item.get(
                    "dataAquisicao",
                    ""
                ),
                "data_baixa": item.get(
                    "_data_baixa",
                    ""
                ),
                "motivo": motivo,
            })

            continue

        # ----------------------------------------------------
        # Duplicado ativo
        # ----------------------------------------------------

        if patrimonio in existentes:

            duplicados += 1

            print(
                "  IGNORADO - patrimônio já cadastrado."
            )

            if IGNORAR_DUPLICADOS:

                continue

        # ----------------------------------------------------
        # Existe na lixeira
        # ----------------------------------------------------

        if patrimonio in lixeira:

            na_lixeira += 1

            print(
                "  IGNORADO - patrimônio já existe na lixeira."
            )

            lista_falhas.append({
                "linha": linha,
                "patrimonio": patrimonio,
                "equipamento": item.get(
                    "nome",
                    ""
                ),
                "marca": item.get(
                    "marca",
                    ""
                ),
                "modelo": item.get(
                    "modelo",
                    ""
                ),
                "local": item.get(
                    "local",
                    ""
                ),
                "responsavel": item.get(
                    "responsavel",
                    ""
                ),
                "categoria": item.get(
                    "categoria",
                    ""
                ),
                "data_aquisicao": item.get(
                    "dataAquisicao",
                    ""
                ),
                "data_baixa": item.get(
                    "_data_baixa",
                    ""
                ),
                "motivo": (
                    "Patrimônio já existe na lixeira."
                ),
            })

            continue

        # ----------------------------------------------------
        # Cadastro
        # ----------------------------------------------------

        ok, motivo, equipamento = cadastrar_equipamento(
            sessao,
            item
        )

        if not ok:

            falhas += 1

            print(
                f"  FALHOU - {motivo}"
            )

            mostrar_item(item)

            lista_falhas.append({
                "linha": linha,
                "patrimonio": item.get(
                    "barcode",
                    ""
                ),
                "equipamento": item.get(
                    "nome",
                    ""
                ),
                "marca": item.get(
                    "marca",
                    ""
                ),
                "modelo": item.get(
                    "modelo",
                    ""
                ),
                "local": item.get(
                    "local",
                    ""
                ),
                "responsavel": item.get(
                    "responsavel",
                    ""
                ),
                "categoria": item.get(
                    "categoria",
                    ""
                ),
                "data_aquisicao": item.get(
                    "dataAquisicao",
                    ""
                ),
                "data_baixa": item.get(
                    "_data_baixa",
                    ""
                ),
                "motivo": motivo,
            })

            continue

        # ----------------------------------------------------
        # Cadastro OK
        # ----------------------------------------------------

        sucessos += 1

        print(
            "  OK - Equipamento cadastrado."
        )

        # Atualiza cache local para evitar duplicação
        # dentro da própria planilha.
        if equipamento:

            novo_barcode = limpar(
                equipamento.get(
                    "barcode",
                    patrimonio
                )
            )

            if novo_barcode:
                existentes[
                    novo_barcode
                ] = equipamento

        # ----------------------------------------------------
        # Dar baixa
        # ----------------------------------------------------

        if item.get("_data_baixa"):

            ok_baixa, motivo_baixa = dar_baixa(
                sessao,
                item,
                equipamento or {}
            )

            if ok_baixa:

                baixas += 1

                print(
                    "  OK - Baixa registrada."
                )

            else:

                print(
                    "  AVISO - Cadastro realizado, "
                    "mas a baixa falhou:"
                )

                print(
                    f"  {motivo_baixa}"
                )

                lista_falhas.append({
                    "linha": linha,
                    "patrimonio": item.get(
                        "barcode",
                        ""
                    ),
                    "equipamento": item.get(
                        "nome",
                        ""
                    ),
                    "marca": item.get(
                        "marca",
                        ""
                    ),
                    "modelo": item.get(
                        "modelo",
                        ""
                    ),
                    "local": item.get(
                        "local",
                        ""
                    ),
                    "responsavel": item.get(
                        "responsavel",
                        ""
                    ),
                    "categoria": item.get(
                        "categoria",
                        ""
                    ),
                    "data_aquisicao": item.get(
                        "dataAquisicao",
                        ""
                    ),
                    "data_baixa": item.get(
                        "_data_baixa",
                        ""
                    ),
                    "motivo": (
                        "Cadastro realizado, "
                        "mas baixa falhou: "
                        + motivo_baixa
                    ),
                })

        time.sleep(PAUSA)

    # --------------------------------------------------------
    # 6. Salvar falhas
    # --------------------------------------------------------

    salvar_falhas(
        lista_falhas
    )

    # --------------------------------------------------------
    # 7. Resumo
    # --------------------------------------------------------

    print()
    print()
    print("=" * 70)
    print("IMPORTAÇÃO FINALIZADA")
    print("=" * 70)

    print(
        f"Total processado : {total}"
    )

    print(
        f"Cadastrados      : {sucessos}"
    )

    print(
        f"Baixas realizadas: {baixas}"
    )

    print(
        f"Duplicados       : {duplicados}"
    )

    print(
        f"Na lixeira       : {na_lixeira}"
    )

    print(
        f"Falhas            : {falhas}"
    )

    print(
        f"Falhas/avisos CSV : {len(lista_falhas)}"
    )

    print("=" * 70)

    if TESTAR_PRIMEIRO:

        print()
        print(
            "O teste foi executado somente para "
            "o primeiro registro."
        )

        print(
            "Se o equipamento apareceu corretamente "
            "no sistema, altere:"
        )

        print(
            "TESTAR_PRIMEIRO = False"
        )

        print(
            "e execute novamente para importar tudo."
        )

    print()

    input(
        "Pressione ENTER para sair..."
    )


# ============================================================
# EXECUÇÃO
# ============================================================

if __name__ == "__main__":

    try:

        main()

    except KeyboardInterrupt:

        print()
        print(
            "Importação interrompida pelo usuário."
        )

    except Exception as erro:

        print()
        print("=" * 70)
        print("ERRO INESPERADO")
        print("=" * 70)
        print(
            f"{type(erro).__name__}: {erro}"
        )
        print()

        input(
            "Pressione ENTER para sair..."
        )